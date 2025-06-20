import time
import os
import sys
import datetime
import numpy as np
import keras
from keras.models import Model
import random
from keras.datasets import mnist
from numpy import arange

# from selection_metrics import *

import argparse

from keras.applications import vgg19,resnet50
from mnist_cifar_imagenet_svhn.tools.utils import NeuronCoverage,TopKNeuronCoverage

basedir = os.path.dirname(os.path.abspath(__file__))

'''
usage: python -u -m mnist_cifar_imagenet_svhn.selection --exp_id=lenet1 
env: pace
'''
threshold = 0

def get_score(x_test, y_test, model):
    #计算准确率
    score = model.evaluate(x_test, y_test, verbose=0)

    # test_result = model.predict(x_test)
    # result = np.argmax(test_result, axis=1)

    # print(result)
    # print(model.metrics_names)
    print('Test loss:', score[0])
    print('Test accuracy:', score[1])
    print('预测错的数目：', len(x_test)*(1-score[1]))
    return score

def get_ds(countlist,res_get_s, sample_size,X_test,Y_test, X_test2, Y_test2, res,ds, my_model,acc):

    #在非异常点中采样
    len_nonnoise = len(X_test) - countlist[0]
    for key in res:
        b = []
        if len(res[key]) > (len_nonnoise/sample_size):
            #mmd方法采样
            for num in range(int(round(len(res[key]) / (len_nonnoise/sample_size)))):
                b.append(res[key][res_get_s[key][num]])
        else:
            b.append(res[key][res_get_s[key][0]])


        # print(len(b))
        for i in range(len(b)):
            X_test2.append(X_test[b[i]])
            Y_test2.append(Y_test[b[i]])
    #新构造的采样测试集
    X_test4 = np.array(X_test2)
    Y_test4 = np.array(Y_test2)

    print(X_test4.shape, Y_test4.shape)
    score = get_score(X_test4, Y_test4, my_model)
    #标准差
    ds.append(np.abs(score[1] - acc))

    coverage1 = TopKNeuronCoverage(model=my_model)
    cov_pace = coverage1.final_coverage(inputs=X_test4,threshold=0.5,K=3)
    print("pace_cov", exp_id, cov_pace)

    coverage2 = TopKNeuronCoverage(model=my_model)
    cov_all = coverage2.final_coverage(inputs=X_test,threshold=0.5,K=3)
    print("all_cov", exp_id, cov_all)

    # coverage1 = TopKNeuronCoverage(model=my_model)
    # cov_pace = coverage1.final_coverage(inputs=X_test4, threshold=0.5, K=3)
    # print("pace_cov", exp_id, cov_pace)


def get_ds_random(countlist,res_get_s, sample_size,X_test,Y_test, X_test2, Y_test2, res,ds, my_model,acc):

    # random.seed(2)

    #在非异常点中采样
    len_nonnoise = len(X_test) - countlist[0]
    for key in res:
        b = []
        if len(res[key]) > (len_nonnoise/sample_size):
            #随机采样
            b = random.sample(res[key], int(round(len(res[key]) / (len_nonnoise/sample_size))))
        else:
            b = random.sample(res[key], 1)

        print(len(b))
        for i in range(len(b)):
            X_test2.append(X_test[b[i]])
            Y_test2.append(Y_test[b[i]])

    X_test4 = np.array(X_test2)
    Y_test4 = np.array(Y_test2)
    print(X_test4.shape, Y_test4.shape)
    score = get_score(X_test4, Y_test4, my_model)
    ds.append(np.abs(score[1] - acc))

def get_std1(X_test, Y_test, a_unoise, countlist,res,label_noise, first_noise,res_get_s,my_model,acc):
    #存储所有标准差
    dss = []

    for j in range(180,181,5):
        ds = []

        X_test2 = []
        Y_test2 = []

        #选择的异常点数目
        len_noise = j*(1-a_unoise)

        print(j)
        #adaptive random
        X_test2.append(X_test[label_noise[first_noise]])
        X_test2.append(X_test[label_noise[np.argmax(dis[first_noise])]])
        Y_test2.append(Y_test[label_noise[first_noise]])                                    
        Y_test2.append(Y_test[label_noise[np.argmax(dis[first_noise])]])

        pre_num = []
        pre_num.append(first_noise)
        pre_num.append(np.argmax(dis[first_noise]))
        while len(X_test2) < len_noise:
            mins = []
            for i in range(len(label_noise)):
                if i not in set(pre_num):
                    min_info = [float('inf'), 0, 0]
                    for l in pre_num:
                        if dis[i][l] < min_info[0]:
                            min_info[0] = dis[i][l]
                            min_info[1] = i
                            min_info[2] = l
                    mins.append(min_info)
            # print(mins,len(pre_num),len(X_test2))
            maxnum = 0
            X_test2.append(X_test[label_noise[mins[0][1]]])
            Y_test2.append(Y_test[label_noise[mins[0][1]]])
            pre_num.append(mins[0][1])
            for i in mins:
                if i[0] > maxnum:
                    X_test2[-1] = X_test[label_noise[i[1]]]
                    Y_test2[-1] = Y_test[label_noise[i[1]]]
                    pre_num[-1] = i[1]
                    # pre_num.append(i[1])

        print("异常点挑选个数：", len(X_test2))
        get_ds(countlist, res_get_s, j*a_unoise, X_test, Y_test, X_test2, Y_test2, res, ds, my_model,acc)

        print(ds)
        ds_mean = np.sqrt(np.mean(np.square(ds), axis=0))
        print(ds_mean)
        dss.append(ds_mean)

    print(dss)
    return dss

def get_std1_random(X_test, Y_test, a_unoise, countlist,res,label_noise, first_noise,res_get_s,my_model,acc):
    dss = []

    for j in range(30,181,5):
        ds = []
        X_test2 = []
        Y_test2 = []
        len_noise = j*(1-a_unoise)
        print(j)

        random_noise = random.sample(label_noise, len_noise)
        for i in range(len(random_noise)):
            X_test2.append(X_test[random_noise[i]])
            Y_test2.append(Y_test[random_noise[i]])

        print("异常点挑选个数：", len(X_test2))
        get_ds(countlist, res_get_s, j*a_unoise, X_test, Y_test, X_test2, Y_test2, res, ds, my_model,acc)
        print(ds)
        ds_mean = np.sqrt(np.mean(np.square(ds), axis=0))
        print(ds_mean)
        dss.append(ds_mean)

    print(dss)
    return dss

def get_mnist(**kwargs):
    (_, _), (X_test, Y_test) = mnist.load_data()
    X_test = X_test.astype('float32')
    X_test = X_test.reshape(X_test.shape[0], 28, 28, 1)
    X_test /= 255
    Y_test = keras.utils.to_categorical(Y_test, 10)
    return X_test,Y_test

def get_mnist_train(**kwargs):
    (X_train, Y_train), (_, _) = mnist.load_data()
    X_train = X_train.astype('float32')
    X_train = X_train.reshape(X_train.shape[0], 28, 28, 1)
    X_train /= 255
    Y_train = keras.utils.to_categorical(Y_train, 10)
    return X_train, Y_train


def get_cifar10(**kwargs):
    from keras.datasets import cifar10
    subtract_pixel_mean = True
    (X_train, Y_train), (X_test, Y_test) = cifar10.load_data()

    # Normalize data.
    X_train = X_train.astype('float32') / 255
    X_test = X_test.astype('float32') / 255

    # If subtract pixel mean is enabled
    if subtract_pixel_mean:
        X_train_mean = np.mean(X_train, axis=0)
        X_train -= X_train_mean
        X_test -= X_train_mean
    Y_test = keras.utils.to_categorical(Y_test, 10)

    return X_test,Y_test


def get_cifar100(**kwargs):
    from keras.datasets import cifar100

    subtract_pixel_mean = True
    # Load the CIFAR10 data.
    (X_train, Y_train), (X_test, Y_test) = cifar100.load_data()

    # Normalize data.
    X_train = X_train.astype('float32') / 255
    X_test = X_test.astype('float32') / 255

    # If subtract pixel mean is enabled
    if subtract_pixel_mean:
        X_train_mean = np.mean(X_train, axis=0)
        X_train -= X_train_mean
        X_test -= X_train_mean

    Y_test = keras.utils.to_categorical(Y_test, 100)
    return X_test,Y_test


def get_imagenet(**kwargs):

    data_path = os.path.join(basedir,'data',"imagenet.npz")
    data = np.load(data_path)
    X_test, Y_test = data['x_test'], data['y_test']
    exp_id = kwargs['exp_id']
    if exp_id == 'vgg19':
        X_test = vgg19.preprocess_input(X_test)
        Y_test = keras.utils.to_categorical(Y_test, num_classes=1000)
    if exp_id == 'resnet50':
        X_test = resnet50.preprocess_input(X_test)
        Y_test = keras.utils.to_categorical(Y_test, num_classes=1000)
    return X_test,Y_test

def get_svhn(**kwargs):
    from mnist_cifar_imagenet_svhn import SVNH_DatasetUtil

    (_, _), (X_test, Y_test) = SVNH_DatasetUtil.load_data()
    return X_test, Y_test

def get_fashion(**kwargs):
    labels_path = os.path.join(basedir,'data','t10k-labels-idx1-ubyte.gz')
    images_path = os.path.join(basedir,'data','t10k-images-idx3-ubyte.gz')
    import gzip
    with gzip.open(labels_path, 'rb') as lbpath:
        labels = np.frombuffer(lbpath.read(), dtype=np.uint8,
                               offset=8)

    with gzip.open(images_path, 'rb') as imgpath:
        images = np.frombuffer(imgpath.read(), dtype=np.uint8,
                               offset=16).reshape(len(labels), 28,28,1)

    X_test = images.astype('float32') / 255.0
    Y_test = keras.utils.to_categorical(labels, 10)
    return X_test, Y_test

def get_catvsdog(**kwargs):
    img_width, img_height = 224,224
    # train_data_dir_dog = 'datas/train/dogs'
    # train_data_dir_cat = 'datas/train/cats'
    validation_data_dir_dog = os.path.join(basedir,'data','validation/dogs')
    validation_data_dir_cat = os.path.join(basedir,'data','validation/cats')
    X_test,Y_test = [],[]
    from PIL import Image
    import numpy as np
    # list = os.listdir(train_data_dir_dog)
    # for i in range(0,len(list)):
    #     path = os.path.join(train_data_dir_dog, list[i])
    #     img = Image.open(path)
    #     img = img.resize((img_width,img_height))
    #     X_train.append(np.asarray(img))
    #     Y_train.append(0)
    #
    # list = os.listdir(train_data_dir_cat)
    # for i in range(0,len(list)):
    #     path = os.path.join(train_data_dir_cat, list[i])
    #     img = Image.open(path)
    #     img = img.resize((img_width, img_height))
    #     X_train.append(np.asarray(img))
    #     Y_train.append(1)

    list = os.listdir(validation_data_dir_dog)
    for i in range(0,len(list)):
        path = os.path.join(validation_data_dir_dog, list[i])
        img = Image.open(path)
        img = img.resize((img_width, img_height))
        X_test.append(np.asarray(img))
        Y_test.append(0)

    list = os.listdir(validation_data_dir_cat)
    for i in range(0,len(list)):
        path = os.path.join(validation_data_dir_cat, list[i])
        img = Image.open(path)
        img = img.resize((img_width, img_height))
        img = np.asarray(img)
        X_test.append(img)
        Y_test.append(1)

    X_test = np.array(X_test).astype('float32')/255.0
    # X_train = np.array(X_train).astype('float32')/255
    Y_test = keras.utils.to_categorical(np.array(Y_test), num_classes=2)

    return X_test,Y_test


def get_traffic_sign(**kwargs):
    data_path = os.path.join(basedir,"data","GTSRB_Test.npz")
    dataset = np.load(data_path)
    x_test, y_test = dataset['x_test'], dataset['y_test']
    x_test = x_test.astype('float32') / 255.0
    y_test = keras.utils.to_categorical(y_test, num_classes=43)
    return x_test,y_test

def get_adv_mnist(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/fgsm_bim_pgd_clean_mnist_image.npy')
    label_path = os.path.join(basedir,'data','adv_image/fgsm_bim_pgd_clean_mnist_label.npy')
    x_test = np.load(image_path)
    # x_test = x_test.astype('float32') / 255.0
    y_test = np.load(label_path)
    #y_test = keras.utils.to_categorical(y_test,num_classes=10)
    return x_test,y_test

def get_label_mnist(**kwargs):
    image_path = os.path.join(basedir,'data','label_shift_data/mnist_label_imgs.npy')
    label_path = os.path.join(basedir,'data','label_shift_data/mnist_label_labels.npy')
    x_test = np.load(image_path)
    # x_test = x_test.astype('float32') / 255.0
    y_test = np.load(label_path)
    #y_test = keras.utils.to_categorical(y_test,num_classes=10)
    return x_test,y_test

def get_corrupted_mnist(**kwargs):
    image_path = os.path.join(basedir,'data','corrupted_image/corrupted_clean_mnist_image.npy')
    label_path = os.path.join(basedir,'data','corrupted_image/corrupted_clean_mnist_label.npy')
    x_test = np.load(image_path)
    y_test = np.load(label_path)
    return x_test,y_test

def get_adv_cifar10(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/bim_cifar10_image.npy')
    label_path = os.path.join(basedir,'data','adv_image/bim_cifar10_label.npy')
    x_test = np.load(image_path).astype('float32')
    y_test = np.load(label_path)
    y_test = keras.utils.to_categorical(y_test,num_classes=10)
    return x_test,y_test

def get_adv_fashion(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/bim_fashion_image.npy')
    label_path = os.path.join(basedir,'data','adv_image/bim_fashion_label.npy')
    x_test = np.load(image_path).astype('float32')
    y_test = np.load(label_path)
    y_test = keras.utils.to_categorical(y_test,num_classes=10)
    return x_test,y_test

def get_adv_svhn(**kwargs):
    image_path = os.path.join(basedir,'data','adv_image/bim_svhn_image.npy')
    label_path = os.path.join(basedir,'data','adv_image/bim_svhn_label.npy')
    x_test = np.load(image_path).astype('float32')
    y_test = np.load(label_path)
    y_test = keras.utils.to_categorical(y_test,num_classes=10)

    return x_test,y_test

def get_combined_fashion(**kwargs):
    image_path = os.path.join(basedir,'data','combined_data/fashion_combined_10000_image.npy')
    label_path = os.path.join(basedir,'data','combined_data/fashion_combined_10000_label.npy')
    x_test = np.load(image_path).astype('float32')
    y_test = np.load(label_path)
    y_test = keras.utils.to_categorical(y_test,num_classes=10)

    return x_test,y_test

def get_combined_cifar10(**kwargs):
    image_path = os.path.join(basedir,'data','combined_data/cifar10_combined_10000_image.npy')
    label_path = os.path.join(basedir,'data','combined_data/cifar10_combined_10000_label.npy')
    x_test = np.load(image_path).astype('float32')
    y_test = np.load(label_path,allow_pickle=True)
    y_test = keras.utils.to_categorical(y_test,num_classes=10)

    return x_test,y_test

def get_combined_svhn(**kwargs):
    image_path = os.path.join(basedir,'data','combined_data/svhn_combined_10000_image.npy')
    label_path = os.path.join(basedir,'data','combined_data/svhn_combined_10000_label.npy')
    x_test = np.load(image_path).astype('float32')
    y_test = np.load(label_path)
    y_test = keras.utils.to_categorical(y_test, num_classes=10)
    return x_test,y_test

def get_mnist_emnist(**kwargs):
    image_path = os.path.join(basedir, 'data', 'natural_shift_data/mnist_emnist_mix_imgs.npy')
    label_path = os.path.join(basedir, 'data', 'natural_shift_data/mnist_emnist_mix_labels.npy')
    x_test = np.load(image_path).astype('float32')
    y_test = np.load(label_path)
    y_test = keras.utils.to_categorical(y_test, num_classes=10)
    # import pdb; pdb.set_trace()
    return x_test, y_test



def get_data(exp_id):
    exp_model_dict = {'lenet1': get_mnist,
                      'lenet4': get_mnist,
                      'lenet5': get_mnist,
                      'mutant1': get_mnist,
                      'mutant2': get_mnist,
                      'mutant3': get_mnist,
                      'cifar10': get_cifar10,
                      'cifar100': get_cifar100,
                      'resnet50':get_imagenet,
                      'vgg19':get_imagenet,
                      'svhn':get_svhn,
                      'fashion':get_fashion,
                      'traffic_sign':get_traffic_sign,
                      'catvsdog':get_catvsdog,
                      'adv_mnist':get_adv_mnist,
                      'adv_cifar10':get_adv_cifar10,
                      'adv_fashion':get_adv_fashion,
                      'adv_svhn':get_adv_svhn,
                      'combined_cifar10':get_combined_cifar10,
                      'combined_fashion':get_combined_fashion,
                      'combined_svhn':get_combined_svhn,
                      'vgg16':get_cifar10,
                      'corrupted_mnist':get_corrupted_mnist,
                      'label_mnist':get_label_mnist,
                      'mnist_emnist':get_mnist_emnist,}

    return exp_model_dict[exp_id](exp_id=exp_id)


def get_model(exp_id):
    basedir = os.path.abspath(os.path.dirname(__file__))

    exp_model_dict = {'lenet1':'model/LeNet-1.h5',
                      'lenet4':'model/LeNet-4.h5',
                      'lenet5':'model/LeNet-5.h5',
                      'mutant1':'model/mutant1.h5',
                      'mutant2':'model/mutant2.h5',
                      'mutant3':'model/mutant3.h5',
                      'cifar10':'model/model_cifar10.h5',
                      'cifar100':'model/model_cifar100.h5',
                      'svhn':'model/model_svhn.hdf5',
                      'fashion':'model/model_fashion.hdf5',
                      'traffic_sign':"model/model_squeezeNet_TSR.hdf5",
                      'adv_mnist':'model/LeNet-5.h5',
                      'adv_cifar10':'model/model_cifar10.h5',
                      'adv_fashion':'model/model_fashion.hdf5',
                      'adv_svhn':'model/model_svhn.hdf5',
                      'combined_cifar10':'model/model_cifar10.h5',
                      'combined_fashion':'model/model_fashion.hdf5',
                      'combined_svhn':'model/model_svhn.hdf5',
                      'vgg16':'model/cifar10-vgg16_model_alllayers.h5',
                      'catvsdog':'model/cats_and_dogs_small_1.h5',
                      'corrupted_mnist':'model/LeNet-5.h5',
                      'label_mnist':'model/LeNet-5.h5',
                      'mnist_emnist':'model/LeNet-5.h5',}

    if exp_id == 'vgg19':
        my_model = vgg19.VGG19(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id == 'resnet50':
        my_model = resnet50.ResNet50(weights='imagenet')
        adam = keras.optimizers.Adagrad(lr=0.01, epsilon=None, decay=0.0)
        my_model.compile(loss='categorical_crossentropy', optimizer=adam, metrics=['accuracy'])
    elif exp_id in exp_model_dict.keys():
        my_model = keras.models.load_model(os.path.join(basedir,exp_model_dict[exp_id]))
    else:
        raise Exception("no such dataset found: {}".format(exp_id))

    return my_model

def get_acc(exp_id):
    acc_dict = {'lenet1':0.9486,
                'lenet4':0.9679,
                'lenet5':0.9868,
                'mutant1':0.7953,
                'mutant2':0.7727,
                'mutant3':0.7914,
                'cifar10':0.9145,
                'cifar100':0.7142,
                'svhn':0.8789566687154272,
                'vgg19':0.6473,
                'resnet50':0.68266,
                'fashion':0.8988,
                'traffic_sign':0.6975455265241488,
                'catvsdog':0.9,
                'adv_mnist':0.9608,
                'adv_cifar10':0.8,
                'adv_fashion':0.8,
                'adv_svhn':0.8,
                'combined_svhn':0.4369,
                'combined_cifar10':0.4292,
                'combined_fashion':0.4531,
                'vgg16.1':0.945739997959137,
                'vgg16':0.9607399998664856}
    return acc_dict[exp_id]

if __name__=="__main__":

    """Parser of command args"""
    parse = argparse.ArgumentParser()
    parse.add_argument("--exp_id", type=str, help="exp_identifiers")
    parse.add_argument("--select_layer_idx", type=int, help="selected feature layer")
    parse.add_argument("--dec_dim", type=int, help='decomposition dim')
    parse.add_argument("--min_samples", type=int,help='min samples cluster')
    parse.add_argument("--min_cluster_size", type=int, help='min_cluster_size')

    console_flags, unparsed = parse.parse_known_args(sys.argv[1:])

    select_layer_idx = console_flags.select_layer_idx
    dec_dim = console_flags.dec_dim
    exp_id = console_flags.exp_id
    min_cluster_size = console_flags.min_cluster_size
    min_samples = console_flags.min_samples
    #acc = get_acc(exp_id)
    # start = datetime.datetime.now()

    my_model = get_model(exp_id=exp_id)
    X_test,Y_test = get_data(exp_id=exp_id)
    print(my_model.summary())
    print(X_test.shape)
    get_score(X_test,Y_test,my_model)
    entropy(X_test, Y_test, my_model, 50)
    start = datetime.datetime.now()

    dense_layer_model = Model(inputs=my_model.input, outputs=my_model.layers[select_layer_idx].output)
    dense_output = dense_layer_model.predict(X_test)
    print(dense_output.shape)

    from sklearn.preprocessing import MinMaxScaler
    import hdbscan

    minMax = MinMaxScaler()
    dense_output = minMax.fit_transform(dense_output)
    print(dense_output)

    if exp_id in ['cifar10','vgg16','lenet1','svhn','vgg19','resnet50','cifar100','traffic_sign','combined_svhn','combined_fashion','combined_cifar10','adv_svhn','adv_fashion','adv_cifar10','catvsdog']:
        from sklearn.decomposition import FastICA
        fica = FastICA(n_components=dec_dim)
        dense_output = fica.fit_transform(dense_output)

    clusterer = hdbscan.HDBSCAN(min_cluster_size=min_cluster_size, min_samples=min_samples)
    r = clusterer.fit(dense_output)

    labels = r.labels_

    print(labels)
    print(np.max(labels))
    print(np.min(labels))

    y_pred_list = labels.tolist()
    countlist = []

    for i in range(np.min(labels), np.max(labels) + 1):
        countlist.append(y_pred_list.count(i))

    print(countlist)

    print(np.sort(countlist))
    print(np.argsort(countlist))

    label_noise = []
    for i, l in enumerate(labels):
        if l == -1:
            label_noise.append(i)

    res = {}
    for i, l in enumerate(labels):
        if l != -1:
            if l not in res:
                res[l] = []
            res[l].append(i)

    print(len(res[0]))

    for key in res:
        X_test3 = []
        Y_test3 = []
        print(len(res[key]))
        for i in range(len(res[key])):
            X_test3.append(X_test[res[key][i]])
            Y_test3.append(Y_test[res[key][i]])
        X_test3 = np.array(X_test3)
        Y_test3 = np.array(Y_test3)
        score = get_score(X_test3, Y_test3, my_model)


    import math

    #计算异常点每两点之间的距离
    dis = np.zeros((len(label_noise), len(label_noise)))
    for i in range(len(label_noise)):
        for j in range(len(label_noise)):
            if j != i:
                dis[i][j] = math.sqrt(np.power(dense_output[label_noise[i]] - dense_output[label_noise[j]], 2).sum())
                # dis[i][j] = math.sqrt(np.power(x_test_de[label_noise[i]] - x_test_de[label_noise[j]], 2).sum())

    noise_score = []
    for i, l in enumerate(r.outlier_scores_):
        if labels[i] == -1:
            noise_score.append(l)
    noise_score = np.array(noise_score)

    # outlier_sort = np.argsort(-r.outlier_scores_)
    first_noise = np.argsort(-noise_score)[0]
    # first_noise = np.argsort(noise_score)[0]
    print(noise_score[first_noise])

    # first_noise = 10

    #非异常点每一类的排序，key类别号
    res_get_s = {}
    from mnist_cifar_imagenet_svhn.mmd_critic.run_digits_new import run
    for key in res:
        temp_dense = []
        for l in res[key]:
            temp_dense.append(dense_output[l])
        temp_dense = np.array(temp_dense)
        temp_label = np.full((len(temp_dense)), key)
        mmd_res, _ = run(temp_dense, temp_label, gamma=0.026, m=min(len(temp_dense), 180), k=0, ktype=0, outfig=None,
                    critoutfig=None, testfile=os.path.join(basedir,'data/a.txt'))
        res_get_s[key] = mmd_res
    import openpyxl
    #将结果保存到excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    row_num = 5
    for a_unoise in arange(0.8,0.9,0.1):
        print(a_unoise)
        dss = get_std1(X_test=X_test, Y_test=Y_test, a_unoise=a_unoise, countlist=countlist, res=res,
                       label_noise=label_noise, first_noise=first_noise, res_get_s=res_get_s, my_model=my_model, acc=acc)

        sheet.cell(row=row_num, column=1).value = a_unoise
        for i in range(len(dss)):
            sheet.cell(row=row_num, column=i + 2).value = dss[i]
        elapsed = (datetime.datetime.now() - start)
        print("Time used: ", elapsed)
        sheet.cell(row=row_num, column=len(dss)+2).value = elapsed
        row_num += 1
    if console_flags.dec_dim is None:
        workbook.save(os.path.join(basedir, "result","{}-sn{}-cs{}.xlsx".format(exp_id, min_samples, min_cluster_size)))
    else:
        workbook.save(os.path.join(basedir, "result","{}-sn{}-cs{}-dim{}.xlsx".format(exp_id, min_samples,
                                                                                          min_cluster_size,
                                                                                          console_flags.dec_dim)))

    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)