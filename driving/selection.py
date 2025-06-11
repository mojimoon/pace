import math
import os
import argparse
from numpy import arange
import random
from driving_models import *
from epoch.epoch_model import *
import numpy as np
from utils import *
import datetime
from data_utils import *
'''
env: pace
usage: python selection.py --exp_id=udacity_dave
'''

def get_score(test_generator, test_samples, batch_size, model):
    results = model.evaluate_generator(test_generator,
                                       steps=math.ceil(test_samples * 1. / batch_size))
    print('mse loss for testing data', results)
    return results


def get_ds(countlist,res_get_s, sample_size,X_test,Y_test, X_test2, Y_test2, res,ds, my_model,acc):

    #在非异常点中采样
    len_nonnoise = len(X_test) - countlist[0]
    for key in res:
        b = []
        if len(res[key]) > (len_nonnoise/sample_size):
            #mmd方法采样
            for num in range(int(round(len(res[key]) / (len_nonnoise/sample_size)))):
                b.append(res[key][res_get_s[key][num]])
        else:
            b.append(res[key][res_get_s[key][0]])


        print(len(b))
        for i in range(len(b)):
            X_test2.append(X_test[b[i]])
            Y_test2.append(Y_test[b[i]])

    X_test4 = np.array(X_test2)
    Y_test4 = np.array(Y_test2)

    print(X_test4.shape, Y_test4.shape)
    score = get_score(X_test4, Y_test4, my_model)
    # ds.append(np.abs(score - acc))
    ds.append(np.abs(score - pre_acc))


def get_ds_random(countlist,res_get_s, sample_size,X_test,Y_test, X_test2, Y_test2, res,ds, my_model,acc):

    #在非异常点中采样
    len_nonnoise = len(X_test) - countlist[0]
    for key in res:
        b = []
        if len(res[key]) > (len_nonnoise/sample_size):
            #随机采样
            b = random.sample(res[key], int(round(len(res[key]) / (len_nonnoise/sample_size))))
        else:
            b = random.sample(res[key], 1)

        print(len(b))
        for i in range(len(b)):
            X_test2.append(X_test[b[i]])
            Y_test2.append(Y_test[b[i]])

    X_test4 = np.array(X_test2)
    Y_test4 = np.array(Y_test2)
    print(X_test4.shape, Y_test4.shape)
    score = get_score(X_test4, Y_test4, my_model)
    ds.append(np.abs(score[1] - acc))


def get_std1(X_test, Y_test, a_unoise, countlist,res,label_noise, first_noise,res_get_s,my_model,acc):
    dss = []

    for j in range(30,181,5):
        ds = []

        X_test2 = []
        Y_test2 = []

        len_noise = j*(1-a_unoise)

        print(j)
        #adaptive random
        X_test2.append(X_test[label_noise[first_noise]])
        X_test2.append(X_test[label_noise[np.argmax(dis[first_noise])]])
        Y_test2.append(Y_test[label_noise[first_noise]])
        Y_test2.append(Y_test[label_noise[np.argmax(dis[first_noise])]])

        pre_num = []
        pre_num.append(first_noise)
        pre_num.append(np.argmax(dis[first_noise]))
        while len(X_test2) < len_noise:
            mins = []
            for i in range(len(label_noise)):
                if i not in set(pre_num):
                    min_info = [float('inf'), 0, 0]
                    for l in pre_num:
                        if dis[i][l] < min_info[0]:
                            min_info[0] = dis[i][l]
                            min_info[1] = i
                            min_info[2] = l
                        mins.append(min_info)
            maxnum = 0
            X_test2.append(X_test[label_noise[mins[0][1]]])
            Y_test2.append(Y_test[label_noise[mins[0][1]]])
            pre_num.append(mins[0][1])
            for i in mins:
                if i[0] > maxnum:
                    X_test2[-1] = X_test[label_noise[i[1]]]
                    Y_test2[-1] = Y_test[label_noise[i[1]]]
                    pre_num[-1] = i[1]

        print("异常点挑选个数：", len(X_test2))
        get_ds(countlist, res_get_s, j*a_unoise, X_test, Y_test, X_test2, Y_test2, res, ds, my_model,acc)

        print(ds)
        ds_mean = np.sqrt(np.mean(np.square(ds), axis=0))
        print(ds_mean)
        dss.append(ds_mean)

    print(dss)
    return dss


def get_std1_random(X_test, Y_test, a_unoise, countlist,res,label_noise, first_noise,res_get_s,my_model,acc):
    dss = []

    for j in range(30,181,5):
        ds = []
        X_test2 = []
        Y_test2 = []
        len_noise = j*(1-a_unoise)
        print(j)

        random_noise = random.sample(label_noise, len_noise)
        for i in range(len(random_noise)):
            X_test2.append(X_test[random_noise[i]])
            Y_test2.append(Y_test[random_noise[i]])

        print("异常点挑选个数：", len(X_test2))
        get_ds(countlist, res_get_s, j*a_unoise, X_test, Y_test, X_test2, Y_test2, res, ds, my_model,acc)
        print(ds)
        ds_mean = np.sqrt(np.mean(np.square(ds), axis=0))
        print(ds_mean)
        dss.append(ds_mean)

    print(dss)
    return dss

basedir = os.path.abspath(os.path.dirname(__file__))
shape = [100,100]
batch_size = 128
def get_udacity_C(**kwargs):
    xs = []
    ys = []

    with open(basedir + '/data' + '/Udacity_C_clean_labeled.txt', 'r') as f:
        for i, line in enumerate(f):
            xs.append(line.split(',')[0])
            ys.append(float(line.split(',')[1]))
    # shuffle list of images
    c = list(zip(xs, ys))
    random.shuffle(c)
    xs, ys = zip(*c)

    train_xs = xs
    train_ys = ys

    train_generator = data_generator(train_xs, train_ys,
                                     target_size=(shape[0], shape[1]),
                                     batch_size=batch_size)

    return train_generator, len(train_xs)

def get_udacity_label(**kwargs):
    xs = []
    ys = []

    with open(basedir + '/data' + '/udacity_label_shifted.txt', 'r') as f:
        for i, line in enumerate(f):
            if i == 0:
                continue
            xs.append(basedir + '/testing' + '/center/' + line.split(',')[0] + '.jpg')
            ys.append(float(line.split(',')[1]))
    # shuffle list of images
    c = list(zip(xs, ys))
    random.shuffle(c)
    xs, ys = zip(*c)

    train_xs = xs
    train_ys = ys

    train_generator = data_generator(train_xs, train_ys,
                                     target_size=(shape[0], shape[1]),
                                     batch_size=batch_size)

    return train_generator, len(train_xs)

def get_udacity_dave(**kwargs):
    xs = []
    ys = []

    with open(basedir + '/data' + '/udacity_dave.txt', 'r') as f:
        for i, line in enumerate(f):
            xs.append(line.split(',')[0])
            ys.append(float(line.split(',')[1]))
    # shuffle list of images
    c = list(zip(xs, ys))
    random.shuffle(c)
    xs, ys = zip(*c)

    train_xs = xs
    train_ys = ys

    train_generator = data_generator(train_xs, train_ys,
                                     target_size=(shape[0], shape[1]),
                                     batch_size=batch_size)

    return train_generator, len(train_xs)

from tensorflow.keras.applications.imagenet_utils import preprocess_input
def get_udacity_adv(**kwargs):
    input_img_data = np.load(basedir + '/data/fgsm_bim_pgd_clean_udacity_eps8_image.npy')
    input_labels = np.load(basedir + '/data/fgsm_bim_pgd_clean_udacity_eps8_label.npy')
    # preprocess img to input to epoch model
    input_img_data = preprocess_input(input_img_data)

    train_generator = data_generator_img(input_img_data, input_labels,
                                     batch_size=batch_size)
    import pdb; pdb.set_trace()
    return train_generator, len(input_labels)

def get_data(exp_id):
    exp_model_dict = {
                      'udacity_C': get_udacity_C,
                      'udacity_label': get_udacity_label,
                      'udacity_adv': get_udacity_adv,
                      'udacity_dave': get_udacity_dave,
                      }

    return exp_model_dict[exp_id](exp_id=exp_id)

def load_data():
    #path = os.path.join(basedir,'testing/final_example.csv')
    print('CH2_final_evaluation.csv contains gt labels. final_example.csv contains false labels.')
    path = os.path.join(basedir, 'testing/CH2_final_evaluation.csv')
    temp = np.loadtxt(path, delimiter=',', dtype=str, skiprows=(1))
    names = list(temp[:, 0])
    test = []
    label = []
    for i in range(len(names)):
        n = names[i]
        path = 'testing/center/' + n + '.jpg'
        path = os.path.join(basedir, path)
        test.append(preprocess_image(path))
        label.append(float(temp[i, 1]))
    test = np.array(test)
    test = test.reshape(test.shape[0], 100, 100, 3)
    label = np.array(label)
    return test, label


def add_black(temp, gradients):
    rect_shape = (30, 30)
    for i in range(temp.shape[0]):
        orig = temp[i].reshape(1, 100, 100, 3)
        grad = gradients[i].reshape(1, 100, 100, 3)
        start_point = (
            random.randint(0, grad.shape[1] - rect_shape[0]), random.randint(0, grad.shape[2] - rect_shape[1]))
        new_grads = np.zeros_like(grad)
        patch = grad[:, start_point[0]:start_point[
            0] + rect_shape[0], start_point[1]:start_point[1] + rect_shape[1]]
        new_grads[:, start_point[0]:start_point[0] + rect_shape[0],
                      start_point[1]:start_point[1] + rect_shape[1]] = -np.ones_like(patch)
        orig = orig + 100 * new_grads
        temp[i] = orig.reshape(100, 100, 3)
    return temp

def get_acc(exp_id):
    #MSE
    acc_dict = {"driving_drop":0.08176111833886715,
                'driving_orig': 0.036451386196547364, #0.09660315929610656,
                'driving_norminit': 0.04329014473285645,
                'black_orig':0.2582445140806614,
                'black_drop':0.2822831075318406,
                'light_orig':0.15019167540836337,
                'light_drop':0.288283072609809}
    return acc_dict[exp_id]

def add_light(temp, gradients):
    temp = temp.reshape(temp.shape[0], -1)
    gradients = gradients.reshape(gradients.shape[0], -1)
    new_grads = np.ones_like(gradients)
    grad_mean = 500 * np.mean(gradients, axis=1)
    grad_mean = np.tile(grad_mean, temp.shape[1])
    grad_mean = grad_mean.reshape(temp.shape)
    temp = temp + 80 * new_grads * grad_mean
    temp = temp.reshape(temp.shape[0], 100, 100, 3)
    return temp

if __name__=="__main__":
    basedir = os.path.abspath(os.path.dirname(__file__))

    """Parser of command args"""
    parse = argparse.ArgumentParser()
    parse.add_argument("--exp_id", type=str, help="exp_identifiers")
    parse.add_argument("--select_layer_idx", type=int, help="selected feature layer")
    parse.add_argument("--dec_dim", type=int, help='decomposition dim')
    parse.add_argument("--min_samples", type=int,help='min samples cluster')
    parse.add_argument("--min_cluster_size", type=int, help='min_cluster_size')

    console_flags, unparsed = parse.parse_known_args(sys.argv[1:])

    select_layer_idx = console_flags.select_layer_idx
    dec_dim = console_flags.dec_dim
    exp_id = console_flags.exp_id
    min_cluster_size = console_flags.min_cluster_size
    min_samples = console_flags.min_samples
    #acc = get_acc(exp_id)

    # input image dimensions
    img_rows, img_cols = 100, 100
    input_shape = (img_rows, img_cols, 3)

    # define input tensor as a placeholder
    input_tensor = Input(shape=input_shape)

    # load multiple models sharing same input tensor
    if exp_id in ['driving_drop','black_drop','light_drop']:
        model = Dave_dropout(input_tensor=input_tensor, load_weights=True)
    elif exp_id in ['driving_orig','black_orig','light_orig']:
        model = Dave_orig(input_tensor=input_tensor, load_weights=True)
    elif exp_id in ['driving_norminit']:
        model = Dave_norminit(input_tensor=input_tensor, load_weights=True)
    elif exp_id in ['udacity_C','udacity_label','udacity_adv','udacity_dave']:
        model = build_cnn() # load epoch model
        model.load_weights("./epoch/epoch.h5")
        K.set_learning_phase(0)

    else:
        raise Exception("no such model {}".format(exp_id))

    print(model.summary())
    # preprocess the data set
    test_generator, test_samples = get_data(exp_id)
    print("data loaded! Num of samples:", test_samples)

    mse = get_score(test_generator, test_samples, batch_size, model)
    print('MSE is ', mse)
    breakpoint()
    start = datetime.datetime.now()

    dense_layer_model = Model(inputs=model.input, outputs=model.layers[select_layer_idx].output)
    dense_output = dense_layer_model.predict(X_test)
    print(dense_output.shape)

    # from sklearn.preprocessing import MinMaxScaler
    #
    # minMax = MinMaxScaler()
    # dense_output = minMax.fit_transform(dense_output)

    from sklearn import preprocessing
    dense_output = preprocessing.normalize(dense_output)

    if exp_id in ['driving_drop','black_drop','light_drop']:
        from sklearn.decomposition import FastICA
        fica = FastICA(n_components=dec_dim)
        dense_output = fica.fit_transform(dense_output)
    import hdbscan
    clusterer = hdbscan.HDBSCAN(min_cluster_size=min_cluster_size, min_samples=min_samples, gen_min_span_tree=True)
    r = clusterer.fit(dense_output)
    labels = r.labels_

    print(labels)
    print(np.max(labels))
    print(np.min(labels))

    y_pred_list = labels.tolist()
    countlist = []

    for i in range(np.min(labels), np.max(labels) + 1):
        countlist.append(y_pred_list.count(i))

    print(countlist)
    print(np.sort(countlist))
    print(np.argsort(countlist))

    label_noise = []
    for i, l in enumerate(labels):
        if l == -1:
            label_noise.append(i)

    res = {}
    for i, l in enumerate(labels):
        if l != -1:
            if l not in res:
                res[l] = []
            res[l].append(i)

    print(len(res[0]))

    for key in res:
        X_test3 = []
        Y_test3 = []
        print(len(res[key]))
        for i in range(len(res[key])):
            X_test3.append(X_test[res[key][i]])
            Y_test3.append(Y_test[res[key][i]])
        X_test3 = np.array(X_test3)
        Y_test3 = np.array(Y_test3)
        score = get_score(X_test3, Y_test3, model)
    import math
    dis = np.zeros((len(label_noise), len(label_noise)))
    for i in range(len(label_noise)):
        for j in range(len(label_noise)):
            if j != i:
                dis[i][j] = math.sqrt(np.power(dense_output[label_noise[i]] - dense_output[label_noise[j]], 2).sum())
                # dis[i][j] = math.sqrt(np.power(x_test_de[label_noise[i]] - x_test_de[label_noise[j]], 2).sum())

    noise_score = []
    for i, l in enumerate(r.outlier_scores_):
        if labels[i] == -1:
            noise_score.append(l)
    noise_score = np.array(noise_score)

    # outlier_sort = np.argsort(-r.outlier_scores_)
    first_noise = np.argsort(-noise_score)[0]
    print(noise_score[first_noise])

    res_get_s = {}
    from driving.mmd_critic.run_digits_new import run
    for key in res:
        temp_dense = []
        for l in res[key]:
            temp_dense.append(dense_output[l])
        temp_dense = np.array(temp_dense)
        temp_label = np.full((len(temp_dense)), key)
        mmd_res, _ = run(temp_dense, temp_label, gamma=0.026, m=min(len(temp_dense), 190), k=0, ktype=0, outfig=None,
                    critoutfig=None, testfile=os.path.join(basedir,'data/a.txt'))
        res_get_s[key] = mmd_res
    import openpyxl
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for a_unoise in arange(0.5, 0.9, 0.1):
        print(a_unoise)
        dss = get_std1(X_test=X_test, Y_test=Y_test, a_unoise=a_unoise, countlist=countlist, res=res,
                       label_noise=label_noise, first_noise=first_noise, res_get_s=res_get_s, my_model=model,acc=acc)

        sheet.cell(row=a_unoise * 10 + 1, column=1).value = a_unoise
        for i in range(len(dss)):
            sheet.cell(row=a_unoise * 10 + 1, column=i + 2).value = dss[i]
        # elapsed = (time.clock() - start)
        # print("Time used: ", elapsed)
        # sheet.cell(row=a_unoise * 10 + 1, column=len(dss)+2).value = elapsed

    #workbook.save(os.path.join(basedir, "result", "{}.xlsx".format(exp_id)))
    if console_flags.dec_dim is None:
        workbook.save(os.path.join(basedir, "result", "{}-sn{}-cs{}.xlsx".format(exp_id, min_samples, min_cluster_size)))
    else:
        workbook.save(os.path.join(basedir, "result", "{}-sn{}-cs{}-dim{}.xlsx".format(exp_id, min_samples,
                                                                                       min_cluster_size,
                                                                                       console_flags.dec_dim)))

    elapsed = (datetime.datetime.now() - start)
    print("Time used: ", elapsed)

